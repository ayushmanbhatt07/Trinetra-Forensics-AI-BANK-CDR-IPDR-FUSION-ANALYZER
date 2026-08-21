import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_styled_excel(columns: tuple | list, rows: list[dict], sheet_title="Export") -> bytes:
    """
    Generates a beautifully styled Excel workbook byte stream.
    Ensures large numeric values (like phone numbers and account numbers) 
    are written as strings to prevent Excel scientific notation (e.g. 9.18E+11).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel limits sheet names to 31 chars
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alt_row_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='B8CCE4'),
        right=Side(style='thin', color='B8CCE4'),
        top=Side(style='thin', color='B8CCE4'),
        bottom=Side(style='thin', color='B8CCE4')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Text coercion heuristics
    text_columns = {"phone", "account", "id", "imsi", "imei", "msisdn", "transaction_id", "txn_id"}
    
    def is_text_column(col_name: str) -> bool:
        col_lower = str(col_name).lower()
        return any(tc in col_lower for tc in text_columns)
    
    # 1. Write Header
    ws.append(columns)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # 2. Write Data
    col_widths = {i: len(str(c)) for i, c in enumerate(columns)}
    
    for row_idx, row_dict in enumerate(rows, start=2):
        row_values = []
        for col_name in columns:
            val = row_dict.get(col_name, "")
            if val is None:
                val = ""
                
            # Coerce numbers to strings if column expects large IDs/phones
            if is_text_column(col_name):
                # We can store as string so excel won't cast to scientific notation
                val = str(val)
                
            row_values.append(val)
            
        ws.append(row_values)
        
        # Apply row styling (alternate shading)
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill
                
            # Update max width for auto-fitting
            col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(cell.value) or ""))
            
    # 3. Auto-fit columns
    for col_idx, max_len in col_widths.items():
        # Add a little padding
        adjusted_width = min(max_len + 2, 50)  # Max width of 50 to avoid absurdly wide cols
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        
    # Freeze the top row
    ws.freeze_panes = "A2"
        
    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
